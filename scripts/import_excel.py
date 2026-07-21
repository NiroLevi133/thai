#!/usr/bin/env python3
"""
ייבוא חד-פעמי מקובץ האקסל אל data/trip.json.
מקור: "מסלול לתאילנד מפורט ביותר!.xlsx" — שני גיליונות:
  - "מלונות שהוזמנו"   → מלונות סגורים (status=booked)
  - "מסלול ב - הנבחר"  → יעדים + מועמדי מלונות (צהוב FFFFFF00 = נסגר)
"""

import json
import os
import re
import sys
from datetime import date, timedelta

import openpyxl

SRC = os.environ.get(
    "TRIP_XLSX",
    os.path.expanduser("~/Downloads/מסלול לתאילנד מפורט ביותר!.xlsx"),
)
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "trip.json")

YELLOW = "FFFFFF00"
SHEET_BOOKED = "מלונות שהוזמנו"
SHEET_ITIN = "מסלול ב - הנבחר"

# הטיול חוצה שנה: כל תאריך לפני 09.12 שייך ל-2027
TRIP_START = date(2026, 12, 9)    # הצ׳ק-אין הראשון בקוסמוי
TRIP_END = date(2027, 1, 14)      # הצ׳ק-אאוט האחרון מבנגקוק

# תאריכי הטיסות בפועל (נמסרו ע"י ניר).
# הלוך: טיסת לילה יוצאת ב-8.12 ונוחתת בתאילנד ב-9.12 (הצ׳ק-אין הראשון).
# חזור: יוצאת אחרי חצות בלילה שבין 13 ל-14.01, ולכן הלילה של ה-13 עדיין במלון.
DEPARTURE = date(2026, 12, 8)
RETURN = date(2027, 1, 14)


# ---------- עזרי פרסינג ----------

def parse_full_date(v):
    """'09.12.2026' → date"""
    if v is None:
        return None
    if isinstance(v, date):
        return v
    m = re.search(r"(\d{1,2})[.\/](\d{1,2})[.\/](\d{2,4})", str(v))
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    return date(y, mo, d)


def parse_short_date(text):
    """'09.12' → date, עם השלמת שנה לפי טווח הטיול"""
    m = re.search(r"(\d{1,2})[.\/](\d{1,2})", text)
    if not m:
        return None
    d, mo = int(m.group(1)), int(m.group(2))
    for y in (2026, 2027):
        try:
            cand = date(y, mo, d)
        except ValueError:
            continue
        if TRIP_START <= cand <= TRIP_END:
            return cand
    return None


def parse_range(text):
    """'09.12 – 13.12' או '29.12 – 03.01' → (start, end)"""
    if not text:
        return None, None
    parts = re.split(r"[–\-—]", str(text))
    if len(parts) < 2:
        return None, None
    return parse_short_date(parts[0]), parse_short_date(parts[1])


def parse_money(v):
    """'2,845 ₪' / '819₪' / '1812 בלי החזר' / '3488 ל 2 לילות' → (מספר, הערה)"""
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip()
    m = re.search(r"[\d,]+(?:\.\d+)?", s)
    if not m:
        return None, s or None
    num = float(m.group(0).replace(",", ""))
    leftover = (s[: m.start()] + " " + s[m.end():]).replace("₪", "").strip()
    leftover = re.sub(r"\s+", " ", leftover)
    return num, leftover or None


def parse_stars(v):
    """'★★★★½' או '⭐⭐⭐⭐' → 4.5 / 4.0"""
    if not v:
        return None
    s = str(v)
    n = s.count("★") + s.count("⭐")
    if not n:
        return None
    if "½" in s:
        n += 0.5
    return float(n)


def strip_stars(name):
    return re.sub(r"[★⭐½\s]+", " ", str(name)).strip() if name else name


def normalize(name):
    """נרמול שם מלון להשוואה בין שני הגיליונות"""
    s = strip_stars(name).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    drop = {
        "hotel", "resort", "spa", "and", "the", "villas", "villa", "pool",
        "sha", "extra", "plus", "adults", "adult", "only", "beach", "koh", "ko",
    }
    return " ".join(w for w in s.split() if w not in drop)


def cell_fill(cell):
    try:
        rgb = cell.fill.start_color.rgb
        return rgb if isinstance(rgb, str) else None
    except Exception:
        return None


def is_yellow(cell):
    return cell_fill(cell) == YELLOW


def hex_color(cell, fallback="#e5e7eb"):
    rgb = cell_fill(cell)
    if not rgb or not isinstance(rgb, str) or len(rgb) != 8 or rgb == "00000000":
        return fallback
    return "#" + rgb[2:]


def iso(d):
    return d.isoformat() if d else None


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")


# ---------- חילוץ מידע מהערות (עמודה J) ----------

EMOJI_RE = re.compile(
    "[\U0001F300-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF\uFE0F\u23F1\u23F0\u2708\u26F4]+"
)


def clean_emoji(text):
    """מנקה אמוג׳י מטקסט חופשי — ב-UI סוג המעבר מיוצג באייקון SVG ובתגית"""
    out = []
    for line in str(text).split("\n"):
        line = EMOJI_RE.sub("", line)
        line = re.sub(r"\s{2,}", " ", line).strip(" +\u00b7-\u2013").strip()
        if line:
            out.append(line)
    return "\n".join(out)


def parse_note(note):
    """מחלץ מס' הזמנה ותאריך ביטול מהערה חופשית בעברית"""
    out = {"confirmationNumber": None, "freeCancelUntil": None}
    if not note:
        return out
    s = str(note)
    m = re.search(r"מס.?\s*הזמנה\s*[-–:]?\s*(\d{6,})", s)
    if m:
        out["confirmationNumber"] = m.group(1)
    m = re.search(r"(?:עד|ה)\s*ה?(\d{1,2}[.\/]\d{1,2}(?:[.\/]\d{2,4})?)", s)
    if m:
        raw = m.group(1)
        d = parse_full_date(raw) if raw.count(".") == 2 else None
        if not d:
            # תאריכי ביטול הם תמיד לפני הטיול (נוב'-דצמ' 2026) או בתוכו
            mm = re.match(r"(\d{1,2})[.\/](\d{1,2})", raw)
            if mm:
                dd, mo = int(mm.group(1)), int(mm.group(2))
                y = 2027 if mo == 1 else 2026
                try:
                    d = date(y, mo, dd)
                except ValueError:
                    d = None
        out["freeCancelUntil"] = iso(d)
    return out


# ---------- קריאת גיליון ההזמנות ----------

def read_booked(wb):
    ws = wb[SHEET_BOOKED]
    rows = []
    for r in range(6, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        total, total_note = parse_money(ws.cell(r, 8).value)
        per_night, _ = parse_money(ws.cell(r, 7).value)
        link_cell = ws.cell(r, 14)
        rows.append({
            "destination": str(ws.cell(r, 1).value or "").strip(),
            "name": str(name).strip(),
            "stars": parse_stars(ws.cell(r, 3).value),
            "checkIn": iso(parse_full_date(ws.cell(r, 4).value)),
            "checkOut": iso(parse_full_date(ws.cell(r, 5).value)),
            "nights": ws.cell(r, 6).value,
            "pricePerNight": per_night,
            "totalPrice": total,
            "bookedVia": ws.cell(r, 9).value,
            "confirmationNumber": str(ws.cell(r, 10).value or "").strip() or None,
            "freeCancelUntil": iso(parse_full_date(ws.cell(r, 11).value)),
            "notes": " ".join(x for x in [ws.cell(r, 13).value, total_note] if x) or None,
            "url": link_cell.hyperlink.target if link_cell.hyperlink else None,
        })
    return rows


# ---------- קריאת גיליון המסלול ----------

def destination_blocks(ws):
    """מחזיר [(start_row, end_row)] לכל בלוק יעד לפי מיזוגים בעמודה A"""
    merged = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col == 1 and rng.min_row >= 2:
            merged[rng.min_row] = rng.max_row

    starts = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            starts.append(r)

    blocks = []
    for i, r in enumerate(starts):
        label = str(ws.cell(r, 1).value).strip()
        if label.startswith("סה"):  # שורת הסיכום "סה"כ לילות - 36"
            continue
        end = merged.get(r, r)
        # הרחבה עד תחילת הבלוק הבא (תופס שורות המשך שלא מוזגו, כמו 19-20, 34-35)
        next_start = next((s for s in starts[i + 1:] if not str(ws.cell(s, 1).value).startswith("סה")), None)
        limit = (next_start - 1) if next_start else ws.max_row
        blocks.append((r, max(end, min(limit, ws.max_row)), label))
    return blocks


def read_itinerary(wb):
    ws = wb[SHEET_ITIN]
    destinations, candidates, transports = [], [], []

    for order, (r0, r1, label) in enumerate(destination_blocks(ws)):
        excel_nights = ws.cell(r0, 2).value
        excel_nights = int(excel_nights) if isinstance(excel_nights, (int, float)) else None
        start, end = parse_range(ws.cell(r0, 3).value)
        # התאריכים הם מקור האמת (הם משתרשרים ברצף); עמודת הלילות באקסל שגויה בפוקט
        nights = (end - start).days if start and end else excel_nights
        dest_id = slug(label)
        destinations.append({
            "id": dest_id,
            "name": label,
            "order": order,
            "startDate": iso(start),
            "endDate": iso(end),
            "nights": nights,
            "nightsConflict": excel_nights if (excel_nights and excel_nights != nights) else None,
            "color": hex_color(ws.cell(r0, 1)),
            "whereToSleep": (str(ws.cell(r0, 5).value).strip() if ws.cell(r0, 5).value else None),
        })

        transfer = ws.cell(r0, 4).value
        if transfer:
            transports.append({"destinationId": dest_id, "description": clean_emoji(transfer), "date": iso(end)})

        # מועמדי מלונות: F/G = מרכזי, H/I = מרוחק
        for name_col, price_col, area in ((6, 7, "central"), (8, 9, "remote")):
            for r in range(r0, r1 + 1):
                cell = ws.cell(r, name_col)
                if not cell.value:
                    continue
                raw = str(cell.value).strip()
                url = cell.hyperlink.target if cell.hyperlink else None
                note = None
                if raw.startswith("http"):
                    url = url or raw
                    raw = "מלון ללא שם"
                    note = "שם חסר באקסל — לעדכן מהלינק"
                price, price_note = parse_money(ws.cell(r, price_col).value)
                row_note = ws.cell(r, 10).value or ws.cell(r0, 10).value if is_yellow(cell) else None
                candidates.append({
                    "destinationId": dest_id,
                    "name": strip_stars(raw),
                    "stars": parse_stars(raw),
                    "area": area,
                    "totalPrice": price,
                    "url": url,
                    "booked": is_yellow(cell),
                    "notes": " · ".join(x for x in [note, price_note] if x) or None,
                    "excelNote": str(row_note).strip() if row_note else None,
                })

    return destinations, candidates, transports


def collect_notes(wb):
    """הערות עמודה J (ממוזגות) לפי שורה"""
    ws = wb[SHEET_ITIN]
    notes = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 10).value
        if v and is_yellow(ws.cell(r, 10)):
            notes.append(str(v).strip())
    return notes


# ---------- הרכבה ----------

def build():
    wb = openpyxl.load_workbook(SRC)
    booked = read_booked(wb)
    destinations, candidates, transfers = read_itinerary(wb)
    excel_notes = collect_notes(wb)

    # מיפוי הערות לפי מספר הזמנה, כדי לזהות סתירות בתאריך ביטול
    note_by_conf = {}
    for n in excel_notes:
        p = parse_note(n)
        if p["confirmationNumber"]:
            note_by_conf[p["confirmationNumber"]] = p

    dest_by_name = {d["name"].lower(): d for d in destinations}

    hotels = []
    used_candidates = set()

    for b in booked:
        dest = dest_by_name.get(b["destination"].lower())
        dest_id = dest["id"] if dest else slug(b["destination"])
        key = normalize(b["name"])

        # מיזוג עם המועמד הצהוב המקביל
        match = None
        for i, c in enumerate(candidates):
            if i in used_candidates or not c["booked"]:
                continue
            if c["destinationId"] == dest_id and (
                normalize(c["name"]) == key
                or key in normalize(c["name"])
                or normalize(c["name"]) in key
            ):
                match = i
                break
        area = "central"
        url = b["url"]
        if match is not None:
            used_candidates.add(match)
            area = candidates[match]["area"]
            url = url or candidates[match]["url"]

        # זיהוי סתירה בתאריך ביטול בין שני המקורות
        conflict = None
        alt = note_by_conf.get(b["confirmationNumber"] or "")
        if alt and alt["freeCancelUntil"] and alt["freeCancelUntil"] != b["freeCancelUntil"]:
            conflict = alt["freeCancelUntil"]

        hotels.append({
            "id": f"h-{slug(b['name'])}",
            "destinationId": dest_id,
            "name": b["name"],
            "stars": b["stars"],
            "area": area,
            "status": "booked",
            "checkIn": b["checkIn"],
            "checkOut": b["checkOut"],
            "nights": b["nights"],
            "pricePerNight": b["pricePerNight"],
            "totalPrice": b["totalPrice"],
            "currency": "ILS",
            "bookedVia": b["bookedVia"],
            "confirmationNumber": b["confirmationNumber"],
            "freeCancelUntil": b["freeCancelUntil"],
            "freeCancelConflict": conflict,
            "paid": False,
            "paidAmount": None,
            "url": url,
            "links": [],
            "notes": b["notes"],
        })

    # מועמדים שנותרו
    seen_ids = set()
    for i, c in enumerate(candidates):
        if i in used_candidates:
            continue
        base = f"c-{c['destinationId']}-{slug(c['name'])}"
        hid, n = base, 2
        while hid in seen_ids:
            hid, n = f"{base}-{n}", n + 1
        seen_ids.add(hid)
        hotels.append({
            "id": hid,
            "destinationId": c["destinationId"],
            "name": c["name"],
            "stars": c["stars"],
            "area": c["area"],
            "status": "candidate",
            "checkIn": None, "checkOut": None, "nights": None,
            "pricePerNight": None,
            "totalPrice": c["totalPrice"],
            "currency": "ILS",
            "bookedVia": None, "confirmationNumber": None,
            "freeCancelUntil": None, "freeCancelConflict": None,
            "paid": False, "paidAmount": None,
            "url": c["url"],
            "links": [],
            "notes": c["notes"],
        })

    # תחבורה: מעברים בין יעדים + טיסות בינ"ל ריקות
    transport = [{
        "id": "t-intl-out",
        "fromDestinationId": None, "toDestinationId": destinations[0]["id"] if destinations else None,
        "date": iso(DEPARTURE), "type": "flight",
        "description": "טיסה מישראל לתאילנד — למלא פרטים",
        "durationMinutes": None, "price": None, "currency": "ILS",
        "status": "idea", "bookingRef": None, "url": None,
    }]
    for i, t in enumerate(transfers):
        desc = t["description"]
        low = desc.lower()
        ttype = "other"
        if "מעבור" in desc or "⛴" in desc:
            ttype = "ferry"
        if "ספיד" in desc:
            ttype = "speedboat"
        if "טיס" in desc or "✈" in desc:
            ttype = "flight"
        if "מיניוואן" in desc or "🚐" in desc:
            ttype = "minivan"
        if "רכבת" in desc:
            ttype = "train"
        elif "אוטובוס" in desc:
            ttype = "bus"
        elif "מונית" in desc:
            ttype = "taxi"
        del low
        idx = next((k for k, d in enumerate(destinations) if d["id"] == t["destinationId"]), None)
        nxt = destinations[idx + 1]["id"] if idx is not None and idx + 1 < len(destinations) else None
        transport.append({
            "id": f"t-{i}-{t['destinationId']}",
            "fromDestinationId": t["destinationId"],
            "toDestinationId": nxt,
            "date": t["date"], "type": ttype,
            "description": desc,
            "durationMinutes": None, "price": None, "currency": "ILS",
            "status": "idea", "bookingRef": None, "url": None,
        })
    transport.append({
        "id": "t-intl-back",
        "fromDestinationId": destinations[-1]["id"] if destinations else None,
        "toDestinationId": None,
        "date": iso(RETURN), "type": "flight",
        "description": "טיסה חזרה לישראל — יוצאת אחרי חצות בלילה שבין 13 ל-14.01",
        "durationMinutes": None, "price": None, "currency": "ILS",
        "status": "idea", "bookingRef": None, "url": None,
    })

    trip = {
        "name": "תאילנד — ניר ואשתו",
        "startDate": iso(TRIP_START),
        "endDate": iso(TRIP_END),
        "departureDate": iso(DEPARTURE),
        "returnDate": iso(RETURN),
        "destinations": destinations,
        "hotels": hotels,
        "transport": transport,
        "attractions": seed_attractions(destinations),
        "expenses": [],
        "settings": {
            "travelers": 2,
            "thbToIls": 0.105,
            "budgetTarget": None,
            "dailyFoodBudget": 120,
            "theme": "light",
        },
    }
    return trip, booked


# ---------- אטרקציות זרע (מינימלי: 2-3 ליעד) ----------

SEED = {
    "koh samui": [("Ang Thong Marine Park — שיט יומי", "טבע", 6.0),
                  ("Big Buddha & Fisherman's Village", "תרבות", 3.0)],
    "ko tao": [("קורס/צלילת היכרות", "ספורט ימי", 4.0),
               ("Nang Yuan Viewpoint", "טבע", 3.0)],
    "koh phangan": [("Bottle Beach + Thong Nai Pan", "חופים", 5.0),
                    ("Phaeng Waterfall Viewpoint", "טבע", 3.0)],
    "krabi": [("Railay Beach + Phra Nang Cave", "חופים", 6.0),
              ("Four Islands Tour", "שיט", 7.0),
              ("Emerald Pool & Hot Springs", "טבע", 5.0)],
    "ko lanta": [("Old Town + חופי המערב", "סיור", 4.0),
                 ("Koh Rok / Koh Haa — שנרקול", "שיט", 8.0)],
    "ko phi phi": [("Maya Bay + Pileh Lagoon", "שיט", 6.0),
                   ("Phi Phi Viewpoint", "טבע", 2.5)],
    "phuket": [("Old Town פוקט", "תרבות", 3.0),
               ("Phang Nga Bay / James Bond Island", "שיט", 8.0),
               ("Promthep Cape — שקיעה", "נוף", 2.0)],
    "khao lak": [("Similan Islands — יום שנרקול/צלילה", "שיט", 9.0),
                 ("Khao Sok National Park", "טבע", 10.0)],
    "pattaya": [("Sanctuary of Truth", "תרבות", 2.5),
                ("Koh Larn — אי סמוך", "חופים", 6.0)],
    "bangkok": [("Grand Palace + Wat Pho", "תרבות", 4.0),
                ("Chatuchak Weekend Market", "שופינג", 4.0),
                ("שוק לילה + רופטופ בר", "בילוי", 4.0)],
}


def seed_attractions(destinations):
    out = []
    for d in destinations:
        for name, cat, hours in SEED.get(d["name"].strip().lower(), []):
            out.append({
                "id": f"a-{d['id']}-{slug(name)}"[:80],
                "destinationId": d["id"],
                "kind": "attraction",
                "name": name,
                "category": cat,
                "price": None,
                "currency": "ILS",
                "durationHours": hours,
                "plannedDate": None,
                "status": "idea",
                "url": None,
                "links": [],
                "notes": None,
            })
    return out


# ---------- אימות ----------

def verify(trip):
    dests = trip["destinations"]
    booked = [h for h in trip["hotels"] if h["status"] == "booked"]
    candidates = [h for h in trip["hotels"] if h["status"] == "candidate"]
    total = sum(h["totalPrice"] or 0 for h in booked)
    nights_planned = sum(d["nights"] or 0 for d in dests)

    covered = set()
    for h in booked:
        if h["checkIn"] and h["checkOut"]:
            a = date.fromisoformat(h["checkIn"])
            b = date.fromisoformat(h["checkOut"])
            while a < b:
                covered.add(a)
                a += timedelta(days=1)

    all_nights = set()
    for d in dests:
        if d["startDate"] and d["endDate"]:
            a = date.fromisoformat(d["startDate"])
            b = date.fromisoformat(d["endDate"])
            while a < b:
                all_nights.add(a)
                a += timedelta(days=1)
    open_nights = sorted(all_nights - covered)

    print(f"  יעדים:            {len(dests)}")
    print(f"  לילות מתוכננים:   {nights_planned}")
    print(f"  מלונות סגורים:    {len(booked)}")
    print(f"  מועמדים:          {len(candidates)}")
    print(f"  סה\"כ שולם/שוריין: {total:,.0f} ₪")
    print(f"  לילות פתוחים:     {len(open_nights)}")
    print(f"  תחבורה:           {len(trip['transport'])} רשומות")
    print(f"  אטרקציות:         {len(trip['attractions'])}")

    errors = []
    if len(dests) != 10:
        errors.append(f"ציפינו ל-10 יעדים, התקבלו {len(dests)}")
    if nights_planned != 36:
        errors.append(f"ציפינו ל-36 לילות, התקבלו {nights_planned}")
    if len(booked) != 8:
        errors.append(f"ציפינו ל-8 מלונות סגורים, התקבלו {len(booked)}")
    if abs(total - 14263) > 5:
        errors.append(f"ציפינו ל-14,263 ₪, התקבל {total:,.0f}")
    return errors, open_nights


if __name__ == "__main__":
    if not os.path.exists(SRC):
        sys.exit(f"❌ קובץ המקור לא נמצא: {SRC}")

    trip, _ = build()
    errors, open_nights = verify(trip)

    if errors:
        print("\n❌ הייבוא נכשל באימות:")
        for e in errors:
            print("   ·", e)
        sys.exit(1)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(trip, f, ensure_ascii=False, indent=2)
    print(f"\n✅ נכתב: {os.path.relpath(OUT)}")
